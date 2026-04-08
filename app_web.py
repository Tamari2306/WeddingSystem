# app.py - corrected version
import os
import logging
import qrcode
import csv
import zipfile
import shutil
import textwrap
import tempfile
from io import BytesIO, StringIO
from datetime import datetime
from functools import wraps
from urllib.parse import quote as url_encode

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    session, jsonify, send_file, make_response, current_app
)
from werkzeug.utils import secure_filename

from dotenv import dotenv_values, load_dotenv

from PIL import Image, ImageDraw, ImageFont  # Pillow
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

from sqlalchemy.sql import func
from sqlalchemy.exc import IntegrityError

# IMPORT THESE FROM MODELS.PY!
# models.py must expose: Guest, init_db, get_db_session
from models import Guest, init_db, get_db_session

# --- Environment Variable Loading Strategy ---
flask_env = os.getenv('FLASK_ENV', 'production')

if flask_env == 'development':
    current_env_file = '.env.development'
    logging.info("Loading environment from .env.development")
else:
    current_env_file = '.env'
    logging.info("Loading environment from .env")

config = {}
if os.path.exists(current_env_file):
    config = dotenv_values(current_env_file)
    for key, value in config.items():
        if value is not None:
            os.environ[key] = value
else:
    logging.warning(f"Environment file {current_env_file} not found; using existing environment variables.")

# --- Configuration for Database ---
DB_FILE = os.getenv("DB_FILE", "guests.db")
DATABASE_URL = f"sqlite:///./{DB_FILE}"

# --- Flask Application Initialization ---
app = Flask(__name__)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
if not app.config['SECRET_KEY']:
    raise ValueError("SECRET_KEY environment variable is not set. Please set a strong, random key in your .env or .env.development file.")

# Provide SQLAlchemy URI to models/init_db if necessary
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL

# --- Folder Configurations ---
UPLOAD_FOLDER = "uploads"
QR_FOLDER_WEB = 'static/qr_codes'
GUEST_CARDS_FOLDER = 'static/guest_cards'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['QR_FOLDER_WEB'] = QR_FOLDER_WEB
app.config['GUEST_CARDS_FOLDER'] = GUEST_CARDS_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(QR_FOLDER_WEB, exist_ok=True)
os.makedirs(GUEST_CARDS_FOLDER, exist_ok=True)

# --- Logging Configuration ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.info(f"Using database: {DATABASE_URL}")

# --- Admin Credentials ---
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "WedSy#01")

# --- Initialize the database using models.py's init_db function ---
with app.app_context():
    init_db(app)

# Utility function for WhatsApp numbers
def to_whatsapp_number(phone):
    phone = str(phone).strip()
    if phone.startswith('+'):
        phone = phone[1:]
    if phone.startswith('0'):
        phone = phone[1:]
    # If number is 9 digits starting with '7' -> Tanzanian local
    if phone.startswith('7') and len(phone) == 9:
        return f"255{phone}"
    return phone

# Jinja globals
app.jinja_env.globals.update(to_whatsapp_number=to_whatsapp_number, url_encode=url_encode)

# --- Decorators ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Please log in first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Helper Functions ---
def get_safe_filename_name_part(name):
    safe_name = (name or "").upper()
    sanitized = "".join(c if c.isalnum() else '_' for c in safe_name)
    return sanitized

def generate_qr_code(data, filename):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(filename)
    return filename

def normalize_card_type(card_type_input, allowed_input=None):
    """
    Determines card_type and allowed based on the rules:
    - S → single (1)
    - D → double (2)
    - If allowed < 2 → single
    - If allowed = 2 → double
    - If allowed >= 3 → family
    """

    # Normalize shortcuts
    if card_type_input:
        card_type = card_type_input.strip().lower()

        if card_type == "s":
            return "single", 1
        if card_type == "d":
            return "double", 2
        if card_type == "family":
            # Family MUST have allowed >= 3
            if allowed_input and int(allowed_input) >= 3:
                return "family", int(allowed_input)
            # If allowed missing or too small → downgrade
            if allowed_input and int(allowed_input) == 2:
                return "double", 2
            return "single", 1

    # If card type not given, infer from allowed
    if allowed_input:
        allowed = int(allowed_input)
        if allowed < 2:
            return "single", 1
        if allowed == 2:
            return "double", 2
        return "family", allowed

    # Fallback default
    return "single", 1


def get_next_visual_id(session):
    """Return next visual_id integer (use provided session)."""
    # session must be an open SQLAlchemy session
    max_id = session.query(func.max(Guest.visual_id)).scalar()
    if max_id is None:
        return 1
    return int(max_id) + 1


# --- Routes ---
@app.route('/')
@login_required
def view_all():
    with get_db_session() as session:
        guests = session.query(Guest).order_by(Guest.visual_id).all()
        # Assign missing visual_id
        missing_id_guests = [g for g in guests if g.visual_id is None]
        for g in missing_id_guests:
            g.visual_id = get_next_visual_id(session)
            session.add(g)
        if missing_id_guests:
            session.commit()
        return render_template('guests.html', guests=guests, current_environment=flask_env)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('username') == ADMIN_USERNAME and request.form.get('password') == ADMIN_PASSWORD:
            session['logged_in'] = True
            flash('Login successful.', 'success')
            return redirect(url_for('view_all'))
        else:
            flash('Invalid credentials.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    session.pop('logged_in', None)
    flash('Logged out.', 'info')
    return redirect(url_for('login'))

# -------------------- add_guest route --------------------
@app.route('/add_guest', methods=['GET', 'POST'])
@login_required
def add_guest():
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        card_type_input = request.form.get('card_type', 'single')
        group_size_input = request.form.get('group_size', '').strip()

        card_type, default_size = normalize_card_type(card_type_input, group_size_input)

        # If family and group_size numeric provided, use it
        if card_type == 'family' and group_size_input.isdigit():
            group_size = int(group_size_input)
        else:
            group_size = default_size

        with get_db_session() as session:
            # Skip duplicates by phone (option 1)
            existing = session.query(Guest).filter_by(phone=phone).first()
            if existing:
                flash(f"Guest with phone {phone} already exists. Skipping duplicate.", "warning")
                return redirect(url_for('add_guest'))

            visual_id = get_next_visual_id(session)
            qr_id = f"GUEST-{visual_id:04d}"
            sanitized_name = get_safe_filename_name_part(name or "GUEST")
            qr_file_path = os.path.join(app.config['QR_FOLDER_WEB'], f"{qr_id}-{sanitized_name}.png")
            try:
                generate_qr_code(qr_id, qr_file_path)
            except Exception as e:
                current_app.logger.warning(f"Could not create QR image: {e}")

            guest = Guest(
                name=name,
                phone=phone,
                qr_code_id=qr_id,
                qr_code_url=f"/static/qr_codes/{qr_id}-{sanitized_name}.png",
                visual_id=visual_id,
                card_type=card_type,
                group_size=group_size,
                checked_in_count=0
            )
            session.add(guest)
            session.commit()
            flash(f"Guest '{name or phone}' added. Card type: {card_type.title()}, allowed entries: {group_size}.", "success")
            return redirect(url_for('view_all'))

    # GET
    return render_template('add_guest.html')

# -------------------- upload_csv route --------------------
@app.route('/upload_csv', methods=['GET', 'POST'])
@login_required
def upload_csv():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash("No file selected.", "danger")
            return redirect(request.url)

        stream = StringIO(file.stream.read().decode("utf-8"))
        reader = csv.DictReader(stream)
        added = 0
        skipped = 0

        def get_row(row, *keys):
            """
            Safely try multiple column names in different cases.
            """
            for k in keys:
                v = row.get(k) or row.get(k.lower()) or row.get(k.capitalize())
                if v:
                    return v.strip()
            return ""

        with get_db_session() as session:
            for row in reader:

                # --------- Read Basic Fields ----------
                name = get_row(row, "name", "Name")
                phone = get_row(row, "phone", "Phone")

                if not phone:
                    skipped += 1
                    continue

                # --------- Normalize Card Type ----------
                raw_type = get_row(row, "Card Type", "card_type", "type")

                def normalize(raw):
                    raw = (raw or "").strip().lower()
                    if raw in ["s", "single"]:
                        return "single"
                    if raw in ["d", "double"]:
                        return "double"
                    if raw in ["f", "family", "group"]:
                        return "family"
                    return "single"  # default

                card_type = normalize(raw_type)

                # --------- Determine Group Size ----------
                if card_type == "single":
                    group_size = 1

                elif card_type == "double":
                    group_size = 2

                else:  # family card
                    allowed_raw = get_row(row, "Allowed", "allowed", "Size", "size", "Group Size", "group_size")
                    try:
                        group_size = max(1, int(allowed_raw))
                    except:
                        group_size = 1

                # --------- Prevent Duplicate Guests ----------
                existing = session.query(Guest).filter_by(phone=phone).first()
                if existing:
                    skipped += 1
                    continue

                # --------- Assign Unique Visual ID ----------
                visual_id = get_next_visual_id(session)
                qr_id = f"GUEST-{visual_id:04d}"

                sanitized_name = get_safe_filename_name_part(name or "GUEST")
                qr_file_path = os.path.join(
                    app.config['QR_FOLDER_WEB'],
                    f"{qr_id}-{sanitized_name}.png"
                )

                # --------- Generate QR Code ----------
                try:
                    generate_qr_code(qr_id, qr_file_path)
                except Exception as e:
                    current_app.logger.warning(f"QR generation failed for {name}: {e}")

                # --------- Create Guest ----------
                guest = Guest(
                    name=name,
                    phone=phone,
                    qr_code_id=qr_id,
                    qr_code_url=f"/static/qr_codes/{qr_id}-{sanitized_name}.png",
                    visual_id=visual_id,
                    card_type=card_type,
                    group_size=group_size,
                    checked_in_count=0
                )

                session.add(guest)
                session.flush()  # Helps with next visual_id
                added += 1

            session.commit()

        flash(f"CSV Processed — Added: {added}, Skipped (duplicates/missing phone): {skipped}", "success")
        return redirect(url_for('view_all'))

    return render_template('upload_csv.html')

# -------------------- update_status route --------------------
@app.route('/update_status', methods=['POST'])
@login_required
def update_status():
    data = request.get_json() or {}
    qr_code_id = data.get("qr_code_id")

    if not qr_code_id:
        return jsonify(success=False, message="Missing qr_code_id.")

    with get_db_session() as session:
        try:
            guest = session.query(Guest).filter_by(qr_code_id=qr_code_id).first()
            if not guest:
                return jsonify(success=False, message="Guest not found.")

            remaining = guest.group_size - guest.checked_in_count

            if remaining <= 0:
                # nothing left
                guest_details = {
                    "visual_id": guest.visual_id,
                    "name": guest.name,
                    "card_type": (guest.card_type or "").title(),
                    "remaining_entries": 0
                }
                return jsonify(success=False, already_entered=True, message="All allowed entries for this card have already checked in.", guest=guest_details)

            # increment checked_in_count by 1 (one person entering)
            guest.checked_in_count = (guest.checked_in_count or 0) + 1

            # if we've used up all entries, set has_entered + entry_time
            if guest.checked_in_count >= guest.group_size:
                guest.has_entered = True
                guest.entry_time = datetime.now()

            session.commit()

            remaining_after = guest.group_size - guest.checked_in_count
            guest_details = {
                "visual_id": guest.visual_id,
                "name": guest.name,
                "card_type": (guest.card_type or "").title(),
                "remaining_entries": remaining_after
            }

            return jsonify(success=True, message="Check-in successful.", guest=guest_details)

        except Exception as e:
            session.rollback()
            current_app.logger.exception(f"Error updating status for {qr_code_id}: {e}")
            return jsonify(success=False, message=f"An error occurred: {e}")


@app.route('/search_guests')
@login_required
def search_guests():
    query = request.args.get('q', '').strip()
    with get_db_session() as session:
        if query:
            guests = session.query(Guest).filter(
                (Guest.name.ilike(f'%{query}%')) | (Guest.phone.ilike(f'%{query}%'))
            ).order_by(Guest.visual_id).all()
        else:
            guests = session.query(Guest).order_by(Guest.visual_id).all()

        # Convert to JSON-friendly dict
        guests_list = []
        for g in guests:
            guests_list.append({
                "visual_id": g.visual_id,
                "name": g.name,
                "phone": g.phone,
                "qr_code_url": g.qr_code_url,
                "has_entered": g.has_entered,
                "entry_time": g.entry_time.strftime('%Y-%m-%d %H:%M:%S') if g.entry_time else 'N/A',
                "card_type": g.card_type
            })
    return jsonify(guests_list)


@app.route('/download_excel')
@login_required
def download_excel():
    with get_db_session() as session:
        guests = session.query(Guest).all()

        # --- Normalize safely ---
        def ct(g): 
            return (g.card_type or "").strip().lower()

        total_guests = len(guests)
        single_cards = sum(1 for g in guests if ct(g) == "single")
        double_cards = sum(1 for g in guests if ct(g) == "double")
        family_cards = sum(1 for g in guests if ct(g) == "family")

        # Sum of allowed = sum(group_size for family cards)
        total_family_allowed = sum(g.group_size for g in guests if ct(g) == "family")

        entered_guests = sum(1 for g in guests if bool(g.has_entered))
        not_entered_guests = total_guests - entered_guests

        # --- Create Workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Guest Report"

        # --- Title ---
        ws["A1"] = "Guest Summary Report"
        ws["A1"].font = Font(size=14, bold=True)

        # --- Summary Section (Top) ---
        summary_data = [
            ("Total Guests", total_guests),
            ("Single Cards", single_cards),
            ("Double Cards", double_cards),
            ("Family Cards", family_cards),
            ("Total Allowed by Family Cards", total_family_allowed),
            ("Guests Entered", entered_guests),
            ("Guests Not Entered", not_entered_guests),
        ]

        row = 3
        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # --- Table Header ---
        table_start = row + 1
        headers = [
            "ID", "Name", "Phone", "QR Code ID",
            "Has Entered", "Entry Time",
            "Visual ID", "Card Type", "Group Size"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=table_start, column=col, value=header)
            cell.font = Font(bold=True)

        # --- Table Rows ---
        for i, g in enumerate(guests, start=table_start + 1):
            ws.cell(i, 1, g.id)
            ws.cell(i, 2, g.name)
            ws.cell(i, 3, g.phone)
            ws.cell(i, 4, g.qr_code_id)
            ws.cell(i, 5, "Entered" if g.has_entered else "Not Entered")
            ws.cell(i, 6, g.entry_time.strftime('%Y-%m-%d %H:%M:%S') if g.entry_time else "")
            ws.cell(i, 7, g.visual_id)
            ws.cell(i, 8, g.card_type)
            ws.cell(i, 9, g.group_size)  # <-- ADDED

        # --- Conditional Formatting for Entered (Column E) ---
        first_data_row = table_start + 1
        last_data_row = table_start + len(guests)

        if last_data_row >= first_data_row:
            entered_range = f"E{first_data_row}:E{last_data_row}"

            ws.conditional_formatting.add(
                entered_range,
                CellIsRule(
                    operator="equal",
                    formula=['"Entered"'],
                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                )
            )

            ws.conditional_formatting.add(
                entered_range,
                CellIsRule(
                    operator="equal",
                    formula=['"Not Entered"'],
                    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                )
            )

        # --- Auto Column Width ---
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # --- Output file ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="guest_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


@app.route('/zip_qr_codes_web')
@login_required
def zip_qr_codes_web():
    memory_file = BytesIO()
    with zipfile.ZipFile(memory_file, 'w') as zf:
        qr_folder = current_app.config.get('QR_FOLDER_WEB', QR_FOLDER_WEB)
        for filename in os.listdir(qr_folder):
            path = os.path.join(qr_folder, filename)
            if os.path.isfile(path):
                zf.write(path, filename)
    memory_file.seek(0)
    return send_file(memory_file, download_name='qr_codes.zip', as_attachment=True, mimetype='application/zip')

@app.route('/edit_guest/<int:guest_id>', methods=['GET', 'POST'])
@login_required
def edit_guest(guest_id):
    with get_db_session() as session:
        try:
            guest = session.get(Guest, guest_id)
            if not guest:
                flash("Guest not found.", "danger")
                return redirect(url_for('view_all'))

            if request.method == 'POST':
                # Update basic fields
                guest.name = request.form.get('name', guest.name).strip()
                guest.phone = to_whatsapp_number(request.form.get('phone', guest.phone))
                guest.has_entered = 'has_entered' in request.form

                # Normalize card type (single/double/family)
                new_card_type_raw = request.form.get('card_type', guest.card_type)
                new_card_type = normalize_card_type(new_card_type_raw)

                # --- FAMILY CARD HANDLING ---
                if new_card_type == "family":
                    size_raw = request.form.get('group_size', '').strip()

                    # Validate group size
                    try:
                        new_group_size = max(1, int(size_raw))
                    except:
                        flash("Invalid group size for family card.", "danger")
                        return redirect(request.url)

                    # Cannot set group size less than people already checked in
                    if new_group_size < guest.checked_in_count:
                        flash(
                            f"Group size cannot be less than the number already checked in ({guest.checked_in_count}).",
                            "danger"
                        )
                        return redirect(request.url)

                # --- SINGLE & DOUBLE AUTOMATIC GROUP SIZE ---
                elif new_card_type == "single":
                    new_group_size = 1

                    if guest.checked_in_count > 1:
                        flash("Guest already used more scans than allowed for a single card.", "danger")
                        return redirect(request.url)

                elif new_card_type == "double":
                    new_group_size = 2

                    if guest.checked_in_count > 2:
                        flash("Guest already used more scans than allowed for a double card.", "danger")
                        return redirect(request.url)

                else:
                    flash("Unknown card type.", "danger")
                    return redirect(request.url)

                # Apply updated card type + group size
                guest.card_type = new_card_type
                guest.group_size = new_group_size

                session.commit()
                flash('Guest updated successfully.', 'success')
                return redirect(url_for('view_all'))

            return render_template('edit_guest.html', guest=guest)

        except Exception as e:
            session.rollback()
            flash(f'Error updating guest: {e}', 'danger')
            current_app.logger.error(f"Error updating guest {guest_id}: {e}", exc_info=True)
            return redirect(url_for('view_all'))


@app.route('/scan_qr')
@login_required
def scan_qr():
    return render_template('scan_qr.html')

@app.route('/delete_guest/<int:guest_id>', methods=['GET'])
@login_required
def delete_guest(guest_id):
    with get_db_session() as session:
        try:
            guest = session.get(Guest, guest_id)
            if not guest:
                flash("Guest not found.", "danger")
                return redirect(url_for('view_all'))

            if guest.qr_code_url:
                qr_file_relative_path = guest.qr_code_url.lstrip('/')
                qr_file_abs_path = os.path.join(current_app.root_path, qr_file_relative_path)
                if os.path.exists(qr_file_abs_path):
                    os.remove(qr_file_abs_path)
                    current_app.logger.info(f"Deleted QR file: {qr_file_abs_path}")
                else:
                    current_app.logger.warning(f"QR file not found for deletion: {qr_file_abs_path}")

            sanitized_name = get_safe_filename_name_part(guest.name)
            guest_card_filename = f"GUEST-{guest.visual_id:04d}-{sanitized_name}.png"
            guest_card_path = os.path.join(current_app.config['GUEST_CARDS_FOLDER'], guest_card_filename)
            if os.path.exists(guest_card_path):
                os.remove(guest_card_path)
                current_app.logger.info(f"Deleted guest card: {guest_card_path}")
            else:
                current_app.logger.warning(f"Guest card not found for deletion: {guest_card_path}")

            session.delete(guest)
            session.commit()
            flash('Guest and associated files deleted.', 'success')
            return redirect(url_for('view_all'))
        except Exception as e:
            session.rollback()
            flash(f'Error deleting guest: {e}', 'danger')
            current_app.logger.error(f"Error deleting guest {guest_id}: {e}", exc_info=True)
            return redirect(url_for('view_all'))

@app.route('/regenerate_qr_codes')
@login_required
def regenerate_qr_codes():
    with get_db_session() as session:
        try:
            guests = session.query(Guest).all()
            qr_folder = current_app.config.get('QR_FOLDER_WEB', QR_FOLDER_WEB)
            for guest in guests:
                if guest.visual_id is None:
                    guest.visual_id = get_next_visual_id(session)
                qr_code_id = f"GUEST-{guest.visual_id:04d}"
                sanitized_name = get_safe_filename_name_part(guest.name)
                qr_file_path = os.path.join(qr_folder, f"{qr_code_id}-{sanitized_name}.png")
                generate_qr_code(qr_code_id, qr_file_path)
                guest.qr_code_id = qr_code_id
                guest.qr_code_url = f"/{os.path.join(qr_folder, f'{qr_code_id}-{sanitized_name}.png')}"
            session.commit()
            flash("QR codes regenerated.", "success")
        except Exception as e:
            session.rollback()
            flash(f"Error regenerating QR codes: {e}", "danger")
            current_app.logger.error(f"Error regenerating QR codes: {e}", exc_info=True)
    return redirect(url_for('view_all'))

@app.route('/generate_guest_cards')
@login_required
def generate_guest_cards():
    CARD_W, CARD_H = 1240, 1748
    NAME_CENTER_Y = 550
    NAME_X = 550
    QR_SIZE = 175
    QR_MARGIN_BOTTOM = 180
    QR_Y = CARD_H - QR_SIZE - QR_MARGIN_BOTTOM
    QR_X = 750
    CARD_TYPE_MARGIN = 45
    CARD_TYPE_Y = CARD_H - CARD_TYPE_MARGIN - 355
    CARD_TYPE_X = 770
    VISUAL_ID_FONT_SIZE = 35
    VISUAL_ID_MARGIN_BOTTOM = 75
    VISUAL_ID_MARGIN_RIGHT = 25

    with get_db_session() as session:
        try:
            template_path = os.path.join("static", "Card Template.jpg")
            output_folder = current_app.config['GUEST_CARDS_FOLDER']

            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Card template not found at: {template_path}.")

            for file in os.listdir(output_folder):
                file_path = os.path.join(output_folder, file)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    flash(f"Could not delete old card {file}: {e}", "warning")
                    current_app.logger.warning(f"Failed to delete old card {file}: {e}")

            guests = session.query(Guest).all()

            font_path = os.path.join("static", "fonts", "Roboto-Bold.ttf")
            if not os.path.exists(font_path):
                raise FileNotFoundError(f"Font file not found at: {font_path}.")

            name_font = ImageFont.truetype(font_path, 50)
            card_type_font = ImageFont.truetype(font_path, 35)
            visual_id_font = ImageFont.truetype(font_path, VISUAL_ID_FONT_SIZE)

            for guest in guests:
                try:
                    img = Image.open(template_path).convert("RGB")
                    draw = ImageDraw.Draw(img)

                    if not guest.qr_code_url:
                        flash(f"No QR URL for guest {guest.name}. Skipping.", "warning")
                        current_app.logger.warning(f"No qr_code_url for guest {guest.name}.")
                        continue

                    qr_file_relative_path = guest.qr_code_url.lstrip('/')
                    qr_file_abs_path = os.path.join(current_app.root_path, qr_file_relative_path)
                    if not os.path.exists(qr_file_abs_path):
                        flash(f"QR code file not found for guest {guest.name} (ID: {guest.visual_id}). Card generation skipped.", "warning")
                        current_app.logger.warning(f"QR code file not found for guest {guest.name}: {qr_file_abs_path}.")
                        continue

                    qr_img = Image.open(qr_file_abs_path).resize((QR_SIZE, QR_SIZE))
                    wrapped_name = textwrap.fill((guest.name or "").upper(), width=20)
                    lines = wrapped_name.split('\n')

                    line_height_estimate = name_font.getbbox("A")[3] + 10
                    total_text_height = line_height_estimate * len(lines)
                    name_y_start = NAME_CENTER_Y - total_text_height // 2

                    for i, line in enumerate(lines):
                        x = NAME_X
                        y = name_y_start + i * line_height_estimate
                        draw.text((x, y), line, font=name_font, fill="#000000")

                    qr_x = QR_X
                    img.paste(qr_img, (qr_x, QR_Y))

                    card_type_text = (guest.card_type or "").upper()
                    draw.text((CARD_TYPE_X, CARD_TYPE_Y), card_type_text, font=card_type_font, fill="#CC3332")

                    visual_id_text = f"NO. {guest.visual_id:04d}"
                    text_bbox = draw.textbbox((0, 0), visual_id_text, font=visual_id_font)
                    text_width = text_bbox[2] - text_bbox[0]
                    text_height = text_bbox[3] - text_bbox[1]
                    visual_id_x = CARD_W - text_width - VISUAL_ID_MARGIN_RIGHT
                    visual_id_y = CARD_H - text_height - VISUAL_ID_MARGIN_BOTTOM
                    draw.text((visual_id_x, visual_id_y), visual_id_text, font=visual_id_font, fill="#CC3332")

                    sanitized_name_part = get_safe_filename_name_part(guest.name)
                    sanitized_filename = f"GUEST-{guest.visual_id:04d}-{sanitized_name_part}.png"
                    img.save(os.path.join(output_folder, sanitized_filename))

                except Exception as loop_e:
                    flash(f"Failed to generate card for {guest.name} (ID: {guest.visual_id}). Error: {str(loop_e)}", "danger")
                    current_app.logger.error(f"Error processing card for guest {guest.visual_id}: {loop_e}", exc_info=True)
                    continue

            flash("Guest invitation cards generated successfully.", "success")

        except FileNotFoundError as fnfe:
            flash(f"Required file not found: {str(fnfe)}", "danger")
            current_app.logger.error(f"File not found during card generation: {fnfe}", exc_info=True)
        except Exception as e:
            flash(f"Error generating cards: {str(e)}", "danger")
            current_app.logger.error(f"Error generating cards: {e}", exc_info=True)

    return redirect(url_for('view_all'))

@app.route('/download_card_by_id/<int:visual_id>')
@login_required
def download_card_by_id(visual_id):
    with get_db_session() as session:
        try:
            guest = session.query(Guest).filter_by(visual_id=visual_id).first()
            if not guest:
                flash("Guest not found.", "danger")
                return redirect(url_for('view_all'))

            template_path = os.path.join("static", "Card Template.jpg")
            font_path = os.path.join("static", "fonts", "Roboto-Bold.ttf")
            qr_file_relative = (guest.qr_code_url or "").lstrip('/')
            qr_file = os.path.join(current_app.root_path, qr_file_relative)

            if not os.path.exists(template_path):
                flash("Card template missing.", "danger")
                return redirect(url_for('view_all'))

            if not os.path.exists(font_path):
                flash("Font file missing.", "danger")
                return redirect(url_for('view_all'))

            if not os.path.exists(qr_file):
                flash(f"QR code missing for {guest.name}.", "danger")
                return redirect(url_for('view_all'))

            img = Image.open(template_path).convert("RGB")
            draw = ImageDraw.Draw(img)
            qr_img = Image.open(qr_file).resize((175, 175))

            name_font = ImageFont.truetype(font_path, 50)
            card_type_font = ImageFont.truetype(font_path, 35)
            visual_id_font = ImageFont.truetype(font_path, 35)

            CARD_W, CARD_H = 1240, 1748
            wrapped = textwrap.fill((guest.name or "").upper(), width=20)
            lines = wrapped.split('\n')
            line_h = name_font.getbbox("A")[3] + 10
            total_h = line_h * len(lines)
            start_y = 550 - total_h // 2
            for i, line in enumerate(lines):
                draw.text((550, start_y + i * line_h), line, font=name_font, fill="#000000")

            img.paste(qr_img, (750, CARD_H - 175 - 180))
            draw.text((770, CARD_H - 45 - 355), (guest.card_type or "").upper(), font=card_type_font, fill="#CC3332")

            vis_text = f"NO. {guest.visual_id:04d}"
            box = draw.textbbox((0,0), vis_text, font=visual_id_font)
            vis_w, vis_h = box[2]-box[0], box[3]-box[1]
            draw.text((CARD_W - vis_w - 25, CARD_H - vis_h - 75), vis_text, font=visual_id_font, fill="#CC3332")

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            img.save(temp_file.name)
            temp_file.close()

            return send_file(temp_file.name, as_attachment=True, download_name=f"Guest-{guest.visual_id:04d}.png")

        except Exception as e:
            flash(f"Error generating card: {e}", "danger")
            current_app.logger.error(f"Error downloading regenerated card: {e}", exc_info=True)
            return redirect(url_for('view_all'))

@app.route('/download_all_cards')
@login_required
def download_all_cards():
    output_folder = current_app.config['GUEST_CARDS_FOLDER']
    if not os.path.exists(output_folder) or not os.listdir(output_folder):
        flash("No invitation cards found. Please generate them first.", "warning")
        return redirect(url_for('view_all'))

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename in os.listdir(output_folder):
            path = os.path.join(output_folder, filename)
            if os.path.isfile(path):
                zip_file.write(path, arcname=filename)

    zip_buffer.seek(0)
    return send_file(zip_buffer, download_name="invitation_cards.zip", as_attachment=True)

@app.route('/guest_report_data')
@login_required
def guest_report_data():
    with get_db_session() as session:
        total_guests = session.query(Guest).count()
        single_cards = session.query(Guest).filter_by(card_type='single').count()
        double_cards = session.query(Guest).filter_by(card_type='double').count()
        family_cards = session.query(Guest).filter_by(card_type='family').count()
        entered_guests = session.query(Guest).filter_by(has_entered=True).count()
        not_entered_guests = total_guests - entered_guests

        data = {
            "total_guests": total_guests,
            "single_cards": single_cards,
            "double_cards": double_cards,
            "family_cards": family_cards,
            "entered_guests": entered_guests,
            "not_entered_guests": not_entered_guests
        }
        return jsonify(data)


@app.route('/guest_report')
@login_required
def guest_report():
    return render_template('guest_report.html')

@app.route('/clear_all_data', methods=['GET'])
@login_required
def clear_all_data():
    with get_db_session() as session:
        try:
            num_deleted_guests = session.query(Guest).delete()
            session.commit()
            flash(f"Successfully deleted {num_deleted_guests} guests from the database.", "success")
            current_app.logger.info(f"Cleared {num_deleted_guests} guests from {DATABASE_URL}")

            qr_folder = current_app.config['QR_FOLDER_WEB']
            for filename in os.listdir(qr_folder):
                file_path = os.path.join(qr_folder, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    current_app.logger.error(f"Error deleting QR file {file_path}: {e}")
                    flash(f"Error deleting QR file {filename}: {e}", "warning")

            guest_cards_folder = current_app.config['GUEST_CARDS_FOLDER']
            for filename in os.listdir(guest_cards_folder):
                file_path = os.path.join(guest_cards_folder, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    current_app.logger.error(f"Error deleting guest card {file_path}: {e}")
                    flash(f"Error deleting guest card {filename}: {e}", "warning")

            return redirect(url_for('view_all'))
        except Exception as e:
            session.rollback()
            flash(f"An error occurred while clearing data: {e}", "danger")
            current_app.logger.error(f"Error clearing all data: {e}", exc_info=True)
            return redirect(url_for('view_all'))
        


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
